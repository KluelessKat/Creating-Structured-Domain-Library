#!/usr/bin/env python3
"""
compare_domain_libraries.py

Compare TSV outputs from different runs of a protein-domain library pipeline
and emit an Excel workbook that makes the differences easy to read.

Two modes
---------

1. PAIRWISE (single comparison):
       python compare_domain_libraries.py pair FILE1.tsv FILE2.tsv \\
           --label1 old --label2 new -o diff.xlsx

2. BATCH (scan two directories for matching step files):
       python compare_domain_libraries.py batch DIR_A DIR_B \\
           --label1 old --label2 new -o pipeline_diff.xlsx

   Files are paired by a "step key" inferred from the filename. By default the
   step key is the filename with any leading "N_" stripped and any trailing
   "_<variant>" stripped; override with --step-pattern if your naming differs.
   Example: 2_domainLibraryStructuredSeq_meta.tsv -> step key
   "domainLibraryStructuredSeq". A file pairs with anything in the other dir
   that yields the same step key.

Matching strategy within a comparison
-------------------------------------
Pairing key:
    (Entry, Domain) if both columns exist, else (Entry) alone.
Within each pairing-key group, when multiple rows share the key (e.g. several
EGF-like repeats in one gene) rows are paired 1-to-1 by closest Start
coordinate using a greedy nearest-neighbor match. Start/End are used to flag
"changed_boundaries" when they exist.

Schema is auto-detected: numeric columns get file1__X / file2__X / delta__X
triplets; non-numeric columns are shown side-by-side without a delta.

Row statuses
------------
  identical           Paired rows agree on every compared column
  changed_values      Paired rows agree on Start/End but a numeric value differs
  changed_boundaries  Paired rows have shifted Start and/or End
  only_in_FILE1       Row present only in file1 (but gene exists in both)
  only_in_FILE2       Row present only in file2 (but gene exists in both)
  gene_only_in_FILE1  Entire gene (Entry) absent from file2
  gene_only_in_FILE2  Entire gene (Entry) absent from file1
"""

from __future__ import annotations
import argparse
import math
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

# Columns that define "which gene/domain are we looking at". Start/End also
# get used for pairing and boundary-shift detection when present.
IDENTITY_CANDIDATES = ['Entry', 'Gene Name', 'Domain']
POSITION_CANDIDATES = ['Start', 'End']

# Tolerance for float equality. Float-valued disorder/aromaticity/etc scores
# almost never match exactly between algorithm versions, so we use a small
# tolerance to separate "genuinely identical" from "basically the same".
FLOAT_TOL = 1e-6

FONT = 'Arial'


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class Schema:
    """Detected column structure for a pair of TSVs being compared."""
    identity_cols: list          # subset of IDENTITY_CANDIDATES present in both
    position_cols: list          # subset of POSITION_CANDIDATES present in both
    numeric_cols: list           # numeric columns present in both (excl identity)
    text_cols: list              # non-numeric, non-identity cols present in both
    key_cols: list               # columns used for pairing-key grouping
    only_in_1: list              # columns unique to file1 (reported but not compared)
    only_in_2: list              # columns unique to file2


@dataclass
class Comparison:
    """Full output of one pairwise comparison, ready to be written to a sheet."""
    step_name: str
    label1: str
    label2: str
    schema: Schema
    diff_df: pd.DataFrame
    n_rows_1: int
    n_rows_2: int
    n_genes_1: int
    n_genes_2: int
    status_counts: dict = field(default_factory=dict)
    numeric_stats: pd.DataFrame = field(default_factory=pd.DataFrame)


# ---------------------------------------------------------------------------
# Schema detection
# ---------------------------------------------------------------------------

def detect_schema(df1: pd.DataFrame, df2: pd.DataFrame) -> Schema:
    cols1, cols2 = set(df1.columns), set(df2.columns)
    shared = cols1 & cols2

    identity_cols = [c for c in IDENTITY_CANDIDATES if c in shared]
    if 'Entry' not in identity_cols:
        sys.exit("ERROR: both files must contain an 'Entry' column.")

    position_cols = [c for c in POSITION_CANDIDATES if c in shared]

    # Decide which shared non-identity columns are numeric by inspecting both
    # frames. A column is numeric only if pandas infers a numeric dtype in
    # both files (prevents a sparse column from being treated as text).
    numeric_cols, text_cols = [], []
    for c in shared - set(identity_cols):
        if pd.api.types.is_numeric_dtype(df1[c]) and pd.api.types.is_numeric_dtype(df2[c]):
            numeric_cols.append(c)
        else:
            text_cols.append(c)

    # Preserve original column order (from file1) for nicer output
    order = {c: i for i, c in enumerate(df1.columns)}
    numeric_cols.sort(key=lambda c: order.get(c, 10**9))
    text_cols.sort(key=lambda c: order.get(c, 10**9))

    # Pairing key: (Entry, Domain) when possible, else (Entry)
    key_cols = ['Entry'] + (['Domain'] if 'Domain' in identity_cols else [])

    only_in_1 = sorted(cols1 - cols2)
    only_in_2 = sorted(cols2 - cols1)

    return Schema(identity_cols, position_cols, numeric_cols, text_cols,
                  key_cols, only_in_1, only_in_2)


# ---------------------------------------------------------------------------
# Pairing
# ---------------------------------------------------------------------------

def pair_group(g1: pd.DataFrame, g2: pd.DataFrame,
               position_cols: list):
    """Greedy nearest-neighbor pairing by Start (if available). Returns
    (matched [(i1,i2)...], unmatched i1 list, unmatched i2 list)."""
    idx1, idx2 = list(g1.index), list(g2.index)

    if 'Start' in position_cols:
        candidates = []
        for i1 in idx1:
            for i2 in idx2:
                d = abs(int(g1.at[i1, 'Start']) - int(g2.at[i2, 'Start']))
                candidates.append((d, i1, i2))
        candidates.sort()
    else:
        # No Start to sort on — pair in row order
        candidates = [(0, i1, i2) for i1, i2 in zip(idx1, idx2)]

    used1, used2, matched = set(), set(), []
    for _, i1, i2 in candidates:
        if i1 in used1 or i2 in used2:
            continue
        matched.append((i1, i2))
        used1.add(i1); used2.add(i2)

    only_1 = [i for i in idx1 if i not in used1]
    only_2 = [i for i in idx2 if i not in used2]
    return matched, only_1, only_2


def classify_pair(r1: pd.Series, r2: pd.Series, schema: Schema) -> str:
    boundary_shift = False
    for c in schema.position_cols:
        if r1[c] != r2[c]:
            boundary_shift = True
            break

    value_diff = False
    for c in schema.numeric_cols:
        if c in schema.position_cols:
            continue
        v1, v2 = r1[c], r2[c]
        if pd.isna(v1) and pd.isna(v2):
            continue
        if pd.isna(v1) or pd.isna(v2) or abs(float(v1) - float(v2)) > FLOAT_TOL:
            value_diff = True
            break

    if not value_diff and not boundary_shift:
        for c in schema.text_cols:
            v1, v2 = r1[c], r2[c]
            if pd.isna(v1) and pd.isna(v2):
                continue
            if v1 != v2:
                value_diff = True
                break

    if boundary_shift:
        return 'changed_boundaries'
    if value_diff:
        return 'changed_values'
    return 'identical'


# ---------------------------------------------------------------------------
# Build the diff DataFrame
# ---------------------------------------------------------------------------

def build_diff(df1: pd.DataFrame, df2: pd.DataFrame, schema: Schema,
               label1: str, label2: str) -> pd.DataFrame:
    def mkrow(status, r1, r2):
        row = {'status': status}
        # Identity columns — take from whichever side is present
        src = r1 if r1 is not None else r2
        for c in schema.identity_cols:
            row[c] = src[c] if c in src else pd.NA

        # Side-by-side numeric columns with delta
        for c in schema.numeric_cols:
            v1 = r1[c] if r1 is not None else pd.NA
            v2 = r2[c] if r2 is not None else pd.NA
            row[f'{label1}__{c}'] = v1
            row[f'{label2}__{c}'] = v2
            try:
                row[f'delta__{c}'] = float(v2) - float(v1)
            except (TypeError, ValueError):
                row[f'delta__{c}'] = pd.NA

        # Side-by-side text columns, no delta
        for c in schema.text_cols:
            row[f'{label1}__{c}'] = r1[c] if r1 is not None else pd.NA
            row[f'{label2}__{c}'] = r2[c] if r2 is not None else pd.NA
        return row

    genes1 = set(df1['Entry'].unique())
    genes2 = set(df2['Entry'].unique())

    groups1 = dict(tuple(g) for g in df1.groupby(schema.key_cols))
    groups2 = dict(tuple(g) for g in df2.groupby(schema.key_cols))

    rows = []
    all_keys = set(groups1) | set(groups2)
    for key in sorted(all_keys, key=lambda k: tuple(str(x) for x in (k if isinstance(k, tuple) else (k,)))):
        entry = key[0] if isinstance(key, tuple) else key
        g1, g2 = groups1.get(key), groups2.get(key)

        if g1 is None:
            status = f'gene_only_in_{label2}' if entry not in genes1 else f'only_in_{label2}'
            for _, r2 in g2.iterrows():
                rows.append(mkrow(status, None, r2))
            continue
        if g2 is None:
            status = f'gene_only_in_{label1}' if entry not in genes2 else f'only_in_{label1}'
            for _, r1 in g1.iterrows():
                rows.append(mkrow(status, r1, None))
            continue

        matched, only_1, only_2 = pair_group(g1, g2, schema.position_cols)
        for i1, i2 in matched:
            rows.append(mkrow(classify_pair(df1.loc[i1], df2.loc[i2], schema),
                              df1.loc[i1], df2.loc[i2]))
        for i1 in only_1:
            rows.append(mkrow(f'only_in_{label1}', df1.loc[i1], None))
        for i2 in only_2:
            rows.append(mkrow(f'only_in_{label2}', None, df2.loc[i2]))

    diff = pd.DataFrame(rows)
    if diff.empty:
        return diff

    # Sort: identical first (matches per-status sheet ordering), then
    # boundary shifts, value shifts, whole-gene differences, per-domain
    # adds/drops.
    status_order = {
        'identical':                   0,
        'changed_boundaries':          1,
        'changed_values':              2,
        f'gene_only_in_{label1}':      3,
        f'gene_only_in_{label2}':      4,
        f'only_in_{label1}':           5,
        f'only_in_{label2}':           6,
    }
    diff['_ord'] = diff['status'].map(status_order).fillna(99)
    diff = diff.sort_values(['_ord'] + schema.identity_cols).drop(columns='_ord')
    return diff.reset_index(drop=True)


def compute_numeric_stats(diff: pd.DataFrame, schema: Schema) -> pd.DataFrame:
    """Per-numeric-column summary of how much each metric shifted among paired rows."""
    if diff.empty:
        return pd.DataFrame()
    paired_mask = diff['status'].isin(['identical', 'changed_values', 'changed_boundaries'])
    paired = diff[paired_mask]
    stats = []
    for c in schema.numeric_cols:
        dcol = f'delta__{c}'
        if dcol not in paired.columns:
            continue
        s = pd.to_numeric(paired[dcol], errors='coerce').dropna()
        if s.empty:
            stats.append({'column': c, 'n_paired': 0,
                          'mean_delta': None, 'median_delta': None,
                          'max_abs_delta': None, 'n_unchanged': 0})
            continue
        stats.append({
            'column': c,
            'n_paired': int(len(s)),
            'mean_delta':    float(s.mean()),
            'median_delta':  float(s.median()),
            'max_abs_delta': float(s.abs().max()),
            'n_unchanged':   int((s.abs() <= FLOAT_TOL).sum()),
        })
    return pd.DataFrame(stats)


# ---------------------------------------------------------------------------
# One pairwise comparison
# ---------------------------------------------------------------------------

def compare_pair(f1: Path, f2: Path, label1: str, label2: str,
                 step_name: str) -> Comparison:
    df1 = pd.read_csv(f1, sep='\t')
    df2 = pd.read_csv(f2, sep='\t')
    schema = detect_schema(df1, df2)
    diff = build_diff(df1, df2, schema, label1, label2)
    counts = diff['status'].value_counts().to_dict() if not diff.empty else {}
    stats = compute_numeric_stats(diff, schema)
    return Comparison(
        step_name=step_name, label1=label1, label2=label2, schema=schema,
        diff_df=diff,
        n_rows_1=len(df1), n_rows_2=len(df2),
        n_genes_1=df1['Entry'].nunique(), n_genes_2=df2['Entry'].nunique(),
        status_counts=counts, numeric_stats=stats,
    )


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill('solid', start_color='305496')   # dark blue
SUBHEADER_FILL = PatternFill('solid', start_color='D9E1F2')  # light blue
HEADER_FONT = Font(name=FONT, bold=True, color='FFFFFF')
SUBHEADER_FONT = Font(name=FONT, bold=True)
DEFAULT_FONT = Font(name=FONT)
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Color scale for delta columns: red (negative) -> white (0) -> blue (positive)
DELTA_RULE = ColorScaleRule(
    start_type='min', start_color='F8696B',
    mid_type='num', mid_value=0, mid_color='FFFFFF',
    end_type='max', end_color='5A8AC6',
)


def _excelify(v):
    """Coerce pandas NA / numpy scalar / non-finite floats into something
    openpyxl can write without complaining."""
    if v is None:
        return None
    if isinstance(v, float) and not math.isfinite(v):
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(v, 'item'):
        try:
            v = v.item()
        except (TypeError, ValueError):
            pass
    return v


def _write_df(ws, df: pd.DataFrame, start_row: int = 1,
              freeze_first_cols: int = 0) -> int:
    """Write a DataFrame into a worksheet starting at start_row. Returns the
    row of the final written row (or start_row if the frame is empty)."""
    if df.empty:
        ws.cell(row=start_row, column=1, value='(no rows)').font = DEFAULT_FONT
        return start_row

    # Header
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=str(col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border = BORDER

    # Body
    for i, (_, record) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=i, column=j, value=_excelify(record[col]))
            cell.font = DEFAULT_FONT
            if isinstance(cell.value, float):
                cell.number_format = '0.0000'

    # Column widths — fit to content (sampled) with a cap
    for j, col in enumerate(df.columns, start=1):
        letter = get_column_letter(j)
        sample_len = 0
        for v in df[col].head(200):
            ev = _excelify(v)
            if ev is not None:
                sample_len = max(sample_len, len(str(ev)))
        width = max(len(str(col)), sample_len) + 2
        ws.column_dimensions[letter].width = min(max(width, 10), 40)

    # Freeze panes (header + first N identity cols)
    if freeze_first_cols > 0:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=freeze_first_cols + 1)
    else:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    last_row = start_row + len(df)

    # Conditional formatting on delta columns
    for j, col in enumerate(df.columns, start=1):
        if isinstance(col, str) and col.startswith('delta__'):
            letter = get_column_letter(j)
            rng = f'{letter}{start_row + 1}:{letter}{last_row}'
            ws.conditional_formatting.add(rng, DELTA_RULE)

    ws.auto_filter.ref = f'A{start_row}:{get_column_letter(len(df.columns))}{last_row}'
    return last_row


def _sanitize_sheet_name(name: str) -> str:
    # Excel sheet names: <=31 chars, no []:*?/\
    clean = re.sub(r'[\[\]:*?/\\]', '_', name)
    return clean[:31]


def _write_status_sheets(wb: Workbook, comp: Comparison, prefix: str = ''):
    """Split the diff into one sheet per status category.

    Ordering: identical first, then boundary shifts, value shifts, whole-gene
    differences, and finally per-domain-only adds/drops. Every sheet is
    created even when empty so users get a predictable tab layout.
    """
    # Freeze the identity columns on each per-status sheet (the status column
    # itself is dropped before writing).
    n_freeze = len(comp.schema.identity_cols)

    statuses = [
        'identical',
        'changed_boundaries',
        'changed_values',
        f'gene_only_in_{comp.label1}',
        f'gene_only_in_{comp.label2}',
        f'only_in_{comp.label1}',
        f'only_in_{comp.label2}',
    ]

    df = comp.diff_df
    for status in statuses:
        sheet_name = _sanitize_sheet_name(f'{prefix}{status}')
        ws = wb.create_sheet(sheet_name)
        if df.empty:
            ws.cell(row=1, column=1, value='(no rows)').font = DEFAULT_FONT
            continue
        sub = df[df['status'] == status]
        if sub.empty:
            ws.cell(row=1, column=1,
                    value=f'(no rows with status "{status}")').font = DEFAULT_FONT
            continue
        sheet_df = sub.drop(columns='status')
        _write_df(ws, sheet_df, freeze_first_cols=n_freeze)


def _write_comparison_summary(ws, comp: Comparison, start_row: int = 1) -> int:
    """Write a header block describing one comparison. Returns next free row."""
    r = start_row
    ws.cell(row=r, column=1, value=f'Step: {comp.step_name}').font = Font(
        name=FONT, bold=True, size=14)
    r += 1
    meta = [
        ('Label 1', comp.label1),
        ('Label 2', comp.label2),
        (f'{comp.label1} rows', comp.n_rows_1),
        (f'{comp.label2} rows', comp.n_rows_2),
        (f'{comp.label1} distinct genes', comp.n_genes_1),
        (f'{comp.label2} distinct genes', comp.n_genes_2),
    ]
    for label, value in meta:
        ws.cell(row=r, column=1, value=label).font = SUBHEADER_FONT
        ws.cell(row=r, column=2, value=_excelify(value)).font = DEFAULT_FONT
        r += 1
    r += 1

    ws.cell(row=r, column=1, value='Row status breakdown').font = SUBHEADER_FONT
    r += 1
    status_order = ['changed_boundaries', 'changed_values',
                    f'gene_only_in_{comp.label1}', f'gene_only_in_{comp.label2}',
                    f'only_in_{comp.label1}', f'only_in_{comp.label2}',
                    'identical']
    for s in status_order:
        count = comp.status_counts.get(s, 0)
        ws.cell(row=r, column=1, value=s).font = DEFAULT_FONT
        ws.cell(row=r, column=2, value=int(count)).font = DEFAULT_FONT
        r += 1
    r += 1

    if comp.schema.only_in_1 or comp.schema.only_in_2:
        ws.cell(row=r, column=1,
                value='Columns unique to one file (not compared)').font = SUBHEADER_FONT
        r += 1
        if comp.schema.only_in_1:
            ws.cell(row=r, column=1, value=f'only in {comp.label1}').font = DEFAULT_FONT
            ws.cell(row=r, column=2, value=', '.join(comp.schema.only_in_1)).font = DEFAULT_FONT
            r += 1
        if comp.schema.only_in_2:
            ws.cell(row=r, column=1, value=f'only in {comp.label2}').font = DEFAULT_FONT
            ws.cell(row=r, column=2, value=', '.join(comp.schema.only_in_2)).font = DEFAULT_FONT
            r += 1
        r += 1

    if not comp.numeric_stats.empty:
        ws.cell(row=r, column=1,
                value='Per-column delta statistics (paired rows only)').font = SUBHEADER_FONT
        r += 1
        r = _write_df(ws, comp.numeric_stats, start_row=r)
        r += 2

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 22
    return r


def write_pairwise_workbook(comp: Comparison, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'
    _write_comparison_summary(ws, comp)

    ws_full = wb.create_sheet('Full_diff')
    _write_df(ws_full, comp.diff_df,
              freeze_first_cols=1 + len(comp.schema.identity_cols))

    _write_status_sheets(wb, comp)
    wb.save(out_path)


def write_batch_workbook(comps, out_path: Path, label1: str, label2: str):
    wb = Workbook()

    # --- Overall summary across all steps ----------------------------------
    ws = wb.active
    ws.title = 'Overall_summary'
    ws.cell(row=1, column=1,
            value=f'Pipeline comparison: {label1} vs {label2}').font = Font(
        name=FONT, bold=True, size=14)

    rollup = []
    for c in comps:
        rollup.append({
            'step': c.step_name,
            f'{label1}_rows': c.n_rows_1,
            f'{label2}_rows': c.n_rows_2,
            f'{label1}_genes': c.n_genes_1,
            f'{label2}_genes': c.n_genes_2,
            'changed_boundaries': c.status_counts.get('changed_boundaries', 0),
            'changed_values':     c.status_counts.get('changed_values', 0),
            f'gene_only_in_{label1}': c.status_counts.get(f'gene_only_in_{label1}', 0),
            f'gene_only_in_{label2}': c.status_counts.get(f'gene_only_in_{label2}', 0),
            f'only_in_{label1}':  c.status_counts.get(f'only_in_{label1}', 0),
            f'only_in_{label2}':  c.status_counts.get(f'only_in_{label2}', 0),
            'identical':          c.status_counts.get('identical', 0),
        })
    rollup_df = pd.DataFrame(rollup)
    _write_df(ws, rollup_df, start_row=3, freeze_first_cols=1)

    # One summary + one full-diff sheet per step
    for c in comps:
        sws = wb.create_sheet(_sanitize_sheet_name(f'sum_{c.step_name}'))
        _write_comparison_summary(sws, c)

    for c in comps:
        dws = wb.create_sheet(_sanitize_sheet_name(f'diff_{c.step_name}'))
        _write_df(dws, c.diff_df,
                  freeze_first_cols=1 + len(c.schema.identity_cols))

    wb.save(out_path)


# ---------------------------------------------------------------------------
# Batch-mode file pairing
# ---------------------------------------------------------------------------

DEFAULT_STEP_RE = re.compile(r'^(?:\d+_)?(?P<step>[^/]+?)(?:_[A-Za-z0-9]+)?\.tsv$')


def step_key(path: Path, pattern: Optional[re.Pattern] = None):
    pat = pattern or DEFAULT_STEP_RE
    m = pat.match(path.name)
    return m.group('step') if m and 'step' in m.groupdict() else None


def find_step_pairs(dir1: Path, dir2: Path,
                    pattern: Optional[re.Pattern] = None):
    def index_dir(d):
        out = {}
        for p in sorted(d.glob('*.tsv')):
            k = step_key(p, pattern)
            if k:
                out.setdefault(k, []).append(p)
        return out

    idx1, idx2 = index_dir(dir1), index_dir(dir2)
    pairs = []
    for k in sorted(set(idx1) & set(idx2)):
        p1, p2 = idx1[k][0], idx2[k][0]
        if len(idx1[k]) > 1 or len(idx2[k]) > 1:
            print(f'NOTE: step key "{k}" matched multiple files; using '
                  f'{p1.name} and {p2.name}', file=sys.stderr)
        pairs.append((k, p1, p2))
    for k in sorted(set(idx1) - set(idx2)):
        print(f'NOTE: no match in {dir2} for step "{k}" ({idx1[k][0].name})',
              file=sys.stderr)
    for k in sorted(set(idx2) - set(idx1)):
        print(f'NOTE: no match in {dir1} for step "{k}" ({idx2[k][0].name})',
              file=sys.stderr)
    return pairs


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _sanitize_label(s: str) -> str:
    return ''.join(ch if ch.isalnum() or ch in '_-' else '_' for ch in s)


def print_console_summary(comp: Comparison):
    print(f'\n=== {comp.step_name}: {comp.label1} vs {comp.label2} ===')
    print(f'  {comp.label1}: {comp.n_rows_1:,} rows / {comp.n_genes_1:,} genes')
    print(f'  {comp.label2}: {comp.n_rows_2:,} rows / {comp.n_genes_2:,} genes')
    statuses = ['changed_boundaries', 'changed_values',
                f'gene_only_in_{comp.label1}', f'gene_only_in_{comp.label2}',
                f'only_in_{comp.label1}', f'only_in_{comp.label2}', 'identical']
    for s in statuses:
        print(f'  {s:32s} {comp.status_counts.get(s, 0):>7,}')
    if comp.schema.only_in_1:
        print(f'  cols only in {comp.label1}: {comp.schema.only_in_1}')
    if comp.schema.only_in_2:
        print(f'  cols only in {comp.label2}: {comp.schema.only_in_2}')


def cmd_pair(args):
    label1 = _sanitize_label(args.label1 or args.file1.stem)
    label2 = _sanitize_label(args.label2 or args.file2.stem)
    if label1 == label2:
        label1 += '_1'; label2 += '_2'
    step = args.step_name or step_key(args.file1) or args.file1.stem
    comp = compare_pair(args.file1, args.file2, label1, label2, step)
    print_console_summary(comp)
    write_pairwise_workbook(comp, args.out)
    print(f'\nWrote: {args.out}')


def cmd_batch(args):
    label1 = _sanitize_label(args.label1 or args.dir1.name)
    label2 = _sanitize_label(args.label2 or args.dir2.name)
    if label1 == label2:
        label1 += '_1'; label2 += '_2'
    pattern = re.compile(args.step_pattern) if args.step_pattern else None
    pairs = find_step_pairs(args.dir1, args.dir2, pattern)
    if not pairs:
        sys.exit('ERROR: no matching step files found between the two directories.')
    comps = []
    for step, p1, p2 in pairs:
        comp = compare_pair(p1, p2, label1, label2, step)
        print_console_summary(comp)
        comps.append(comp)
    write_batch_workbook(comps, args.out, label1, label2)
    print(f'\nWrote: {args.out}')


def build_parser():
    p = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = p.add_subparsers(dest='mode', required=True)

    pp = sub.add_parser('pair', help='Compare two TSV files')
    pp.add_argument('file1', type=Path)
    pp.add_argument('file2', type=Path)
    pp.add_argument('--label1', default=None)
    pp.add_argument('--label2', default=None)
    pp.add_argument('--step-name', default=None,
                    help='Label for this step (default: inferred from filename)')
    pp.add_argument('-o', '--out', type=Path, default=Path('diff.xlsx'))
    pp.set_defaults(func=cmd_pair)

    bp = sub.add_parser('batch', help='Compare matching step files in two directories')
    bp.add_argument('dir1', type=Path)
    bp.add_argument('dir2', type=Path)
    bp.add_argument('--label1', default=None)
    bp.add_argument('--label2', default=None)
    bp.add_argument('--step-pattern', default=None,
                    help=r'Regex with a (?P<step>...) group. Default strips leading "N_" and trailing "_variant".')
    bp.add_argument('-o', '--out', type=Path, default=Path('pipeline_diff.xlsx'))
    bp.set_defaults(func=cmd_batch)

    return p


def main():
    args = build_parser().parse_args()
    args.func(args)


if __name__ == '__main__':
    main()
